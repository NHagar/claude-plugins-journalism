"""
Sportförderung in Deutschland – Haushalte und Etats
Bundesministerien, Landesministerien und Kommunen
Stand: 2024/2025 (ergänzt 2026/2027 wo verfügbar)

Primärquellen:
- Bundeshaushalt BMI Einzelplan 06, Kapitel 0602, Titelgruppe 02
- Bundeshaushalt ab 2026: Bundeskanzleramt Einzelplan 04
- Bundestag Drucksachen und hib-Meldungen
- DOSB Übersicht Bundeshaushalt 2025/2026
- Bayerisches StMI / Bayerischer Haushaltsplan Epl. 03, Kap. 03 03
- Berliner Haushalt Epl. 05, Kap. 0510, Tit. 68419
- Landeshaushalte aller 16 Bundesländer (Landessportbünde, Ministerien)
- Kommunale Haushalte (München, Berlin, Köln, Hamburg u.a.)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Datenbasis ────────────────────────────────────────────────────────────────
# Spalten:
#  0 Ebene          (Bund / Land / Kommune)
#  1 Bundesland     (z.B. "Bayern" oder "Gesamt (Bund)")
#  2 Träger         (Ministerium / Behörde)
#  3 Haushaltsstelle(Einzelplan, Kapitel, Titel-Nr.)
#  4 Maßnahme       (Kurztitel)
#  5 Beschreibung   (Langtext)
#  6 Förderbereich  (Spitzensport / Breitensport / Spitzensport + Breitensport)
#  7 Betrag_Mio_EUR (float)
#  8 Jahr           (int)
#  9 Quelle         (URL / Anmerkung)

DATA = [

    # ══════════════════════════════════════════════════════════
    # BUNDESEBENE
    # ══════════════════════════════════════════════════════════

    # ── Gesamtetat BMI 2024 ────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Gesamtsportetat BMI 2024",
     "Gesamter Sportetat des BMI (Titelgruppe 02) – alle Sparten inkl. Spitzensport, "
     "Verbände, Infrastruktur, Safe Sport, NADA",
     "Spitzensport + Breitensport", 282.55, 2024,
     "Bundestag hib; bmi.bund.de – Sportetat 2024"),

    # ── Zentrale Maßnahmen Sport ──────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Zentrale Maßnahmen Sport (Gesamtposition)",
     "Zentrale Maßnahmen auf dem Gebiet des Sports – Spitzenverbände, Kader, "
     "Trainerförderung, Olympiastützpunkte (2024-Ansatz)",
     "Spitzensport", 177.88, 2024,
     "bmi.bund.de; Aufschlüsselung BMI-Übersicht PDF Bundestag 15.10.2024"),

    # ── Olympia-/Perspektivkader ──────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02 – TA Kader",
     "Olympia-/Perspektivkader & int. Wettkämpfe",
     "Förderung olympischer Kader, Perspektivkader und Vorbereitung auf "
     "Weltmeisterschaften und internationale Wettkämpfe",
     "Spitzensport", 50.32, 2025,
     "Bundestag hib 2025 (Planansatz 2025); dosb.de Übersicht BH 2026"),

    # ── Leistungssportpersonal ────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02 – TA Personal",
     "Leistungssportpersonal (Trainer, Management)",
     "Mischfinanzierte Bundestrainer und Managementpersonal der Bundesfachverbände; "
     "für 2025–2028 jährlich rund 39 Mio. €",
     "Spitzensport", 58.46, 2025,
     "BMI Förderentscheidung Sept. 2024; bmi.bund.de Athletenförderung"),

    # ── Olympiastützpunkte ────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02 – TA OSP",
     "Olympiastützpunkte und Trainingszentren",
     "Betrieb (bis 24 Mio. €/Jahr) und Investitionen der Olympiastützpunkte (OSP) "
     "sowie Bundesleistungszentren",
     "Spitzensport", 58.10, 2025,
     "Bundestag hib 2025 (Planansatz 2025)"),

    # ── Sportgroßveranstaltungen ──────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, Tit. 683 xx",
     "Sportgroßveranstaltungen",
     "Beteiligung des Bundes an der Ausrichtung internationaler Sportgroßveranstaltungen "
     "in Deutschland (2025: 44,54 Mio. € u.a. Olympia-Bewerbung)",
     "Spitzensport", 7.31, 2024,
     "BMI Förderübersicht 2024; Bundestag hib"),

    # ── Nicht-olympische Verbände ─────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, Tit. 684 xx",
     "Nicht-olympische Verbände & World Games",
     "Förderung von Verbänden nicht-olympischer Sportarten inkl. World Games",
     "Spitzensport", 13.50, 2024,
     "BMI Förderübersicht 2024"),

    # ── IAT / FES ─────────────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "IAT & FES – Sportwissenschaft",
     "Institut für Angewandte Trainingswissenschaften (IAT, Leipzig) und "
     "Institut für Forschung und Entwicklung von Sportgeräten (FES, Berlin)",
     "Spitzensport", 7.09, 2024,
     "BMI Förderübersicht 2024"),

    # ── NADA ──────────────────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "NADA – Nationale Anti-Doping-Agentur",
     "Zuschuss Bund an die Nationale Anti Doping Agentur Deutschland (NADA)",
     "Spitzensport", 10.38, 2024,
     "BMI Förderübersicht 2024"),

    # ── WADA ──────────────────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "WADA – Beitrag Deutschland",
     "Pflichtbeitrag Deutschlands an die World Anti-Doping Agency (WADA)",
     "Spitzensport", 1.26, 2024,
     "BMI Förderübersicht 2024"),

    # ── Safe Sport ────────────────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Zentrum Safe Sport",
     "Unabhängige Anlauf- und Beratungsstelle für Betroffene von Missbrauch im Sport",
     "Spitzensport + Breitensport", 1.25, 2024,
     "BMI Förderübersicht 2024"),

    # ── Spitzensport-Agentur ──────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Spitzensport-Agentur (Aufbau)",
     "Anschubfinanzierung für die neue Nationale Spitzensport-Agentur",
     "Spitzensport", 0.20, 2024,
     "BMI Förderübersicht 2024"),

    # ── Sportstätten-Infrastruktur ────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02 – Bau",
     "Sportstätten Höchstleistungssport (Bauunterhaltung)",
     "Unterrichtung, Ausstattung und Bauunterhaltung von Sportstätten des Bundes "
     "für den Höchstleistungssport (Vorjahr: 24,6 Mio. €)",
     "Spitzensport", 18.82, 2024,
     "BMI Förderübersicht 2024; bmi.bund.de Infrastrukturförderung"),

    # ── Jahresplanung Verbände 2024 ───────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Jahresplanung Spitzenverbände 2024",
     "Jahresplanung Trainings- und Lehrgangsmaßnahmen der Spitzenverbände (Herbstplanung Dez. 2024)",
     "Spitzensport", 41.00, 2024,
     "BMI Pressemitteilung Dez. 2024"),

    # ── Bundeshaushalt 2025 Gesamtetat ───────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium des Innern und für Heimat (BMI)",
     "EP 06, Kap. 0602, TGr. 02",
     "Gesamtsportetat BMI 2025",
     "Gesamter Sportetat des BMI 2025 – Steigerung gegenüber 2024 um rund 50 Mio. €",
     "Spitzensport + Breitensport", 333.00, 2025,
     "Bundestag hib Haushalt 2025 – Mehr Geld für den Sport"),

    # ── Bundeshaushalt 2026 ───────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundeskanzleramt – Staatsministerin für Sport und Ehrenamt",
     "EP 04 (Bundeskanzleramt)",
     "Gesamtsportetat Bundeskanzleramt 2026 (Rekord)",
     "Ab 2026 Zuständigkeit im EP 04 (Bundeskanzleramt); Zentrale Maßnahmen Sport: "
     "222 Mio. €, Infrastruktur: 48,1 Mio. €, Sondervermögen kommunale Sportstätten: "
     "333 Mio. € + Schwimmbäder 250 Mio. €",
     "Spitzensport + Breitensport", 357.50, 2026,
     "Bundestag hib 1111704; dosb.de Bundeshaushalt 2026"),

    # ── Sondervermögen kommunale Sportstätten ─────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium für Wohnen, Stadtentwicklung und Bauwesen (BMWSB)",
     "EP 25, Sondervermögen Infrastruktur – Programm Sportstätten",
     "Sanierung kommunaler Sportstätten (Tranche 1)",
     "Neues Bundesprogramm 'Sanierung kommunaler Sportstätten' aus dem Sondervermögen "
     "Infrastruktur; für Q1-Aufruf Antragssumme >7,5 Mrd. €",
     "Breitensport", 333.00, 2025,
     "BMWSB Pressemitteilung 2025; bisp-sportinfrastruktur.de"),

    # ── Investitionspakt Sportstätten ─────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium für Wohnen, Stadtentwicklung und Bauwesen (BMWSB)",
     "EP 25, Investitionspakt Sportstätten",
     "Investitionspakt kommunale Sportstätten 2024",
     "Bundesanteil am Investitionspakt zur Förderung kommunaler Sportstätten "
     "(ergänzende Förderung neben Länder- und Kommunalmitteln)",
     "Breitensport", 60.50, 2024,
     "BMWSB; bisp-sportinfrastruktur.de – 276 Mio. € gesamt"),

    # ── Sanierung kommunaler Einrichtungen ───────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium für Wohnen, Stadtentwicklung und Bauwesen (BMWSB)",
     "EP 25, Sondervermögen KTF",
     "Sanierung kommunaler Einrichtungen (Sport, Jugend, Kultur)",
     "Bundesprogramm aus dem Klima- und Transformationsfonds; "
     "476 Mio. € im Wirtschaftsplan; davon 200 Mio. € neu 2024",
     "Breitensport", 364.60, 2024,
     "Bundestag Drucksache 20/14971"),

    # ── Kinder- und Jugendplan ────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium für Familie, Senioren, Frauen und Jugend (BMFSFJ)",
     "EP 17, Kinder- und Jugendplan (KJP)",
     "Kinder- und Jugendplan – Sport / Breitensport",
     "Förderung von Jugend- und Breitensport durch Sportverbände im Rahmen des KJP "
     "(Kürzung gegenüber 2023: 239,1 Mio. €)",
     "Breitensport", 194.50, 2024,
     "Bundestag; BMFSFJ KJP 2024"),

    # ── Bundeswehr Spitzensportförderung ──────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundesministerium der Verteidigung (BMVg) – Bundeswehr",
     "EP 14 (Verteidigung) – Personalkosten Sportsoldaten",
     "Bundeswehr-Spitzensportförderung (890 Dienstposten)",
     "Größter staatlicher Einzelförderer des deutschen Spitzensports; "
     "14 Sportfördergruppen bundesweit, über 220 Disziplinen in >66 Verbänden; "
     "890 Dienstposten für Sportsoldatinnen und -soldaten",
     "Spitzensport", 67.00, 2025,
     "bundeswehr.de – Spitzensport Sportförderer Bundeswehr (Stand März 2025)"),

    # ── Olympia-Bewerbung 2036 ────────────────────────────────
    ("Bund", "Gesamt (Bund)",
     "Bundeskanzleramt – Staatsministerin für Sport und Ehrenamt",
     "EP 04, TGr. 02",
     "Olympia-Bewerbung Deutschland (LA 2036)",
     "Bundesbeteiligung an der deutschen Olympia-Bewerbung für 2036",
     "Spitzensport", 0.91, 2026,
     "Bundestag hib 1111704; Das Parlament – Rekordwert Richtung Olympia"),

    # ══════════════════════════════════════════════════════════
    # LANDESEBENE
    # ══════════════════════════════════════════════════════════

    # ── Bayern ────────────────────────────────────────────────
    ("Land", "Bayern",
     "Bayerisches Staatsministerium des Innern, für Sport und Integration (StMI)",
     "EP 03, Kap. 03 03, Tit. 684 91",
     "Zuschüsse Breiten- und Nachwuchsleistungssport",
     "Institutionelle Förderung Sportverbände und -dachorganisationen "
     "(Breiten-, Nachwuchsleistungssport, Sondermaßnahmen)",
     "Spitzensport + Breitensport", 38.31, 2024,
     "Bayerischer Haushaltsplan Epl. 03, Kap. 03 03; blsv.de – Rekordniveau 2024"),

    ("Land", "Bayern",
     "Bayerisches Staatsministerium des Innern, für Sport und Integration (StMI)",
     "EP 03, Kap. 03 03 – Gesamtvolumen",
     "Gesamte Sportförderung Bayern 2024",
     "Breiten- und Nachwuchsleistungssportförderung inkl. Vereinspauschale (+10 Mio. €), "
     "Verbandsförderung (+10 Mio. €), Seepferdchen-Gutschein (10,8 Mio. €), "
     "Sondermaßnahmen (4,1 Mio. €)",
     "Spitzensport + Breitensport", 110.60, 2024,
     "StMI Bayern Pressemitteilung 2024; blsv.de – Sportförderung auf Rekordniveau"),

    ("Land", "Bayern",
     "Bayerisches Staatsministerium des Innern, für Sport und Integration (StMI)",
     "EP 03, Kap. 03 03",
     "Seepferdchen-Gutscheinprogramm",
     "Staatlich gefördertes Schwimmkurs-Gutscheinprogramm für Kinder in Bayern",
     "Breitensport", 10.80, 2024,
     "StMI Bayern; Bayerischer Haushaltsplan 2024"),

    ("Land", "Bayern",
     "Bayerisches Staatsministerium des Innern, für Sport und Integration (StMI)",
     "EP 03, Kap. 03 03",
     "Sonderfördermaßnahmen Sport Bayern",
     "Besondere Sportfördermaßnahmen: Integration, Inklusion, Sportprojekte",
     "Breitensport", 4.10, 2024,
     "StMI Bayern Pressemitteilung 2024"),

    # ── Nordrhein-Westfalen ───────────────────────────────────
    ("Land", "Nordrhein-Westfalen",
     "Staatskanzlei NRW – Sport und Ehrenamt / MIK NRW",
     "NRW-Haushalt – Sportpauschale (GFG)",
     "Sportpauschale NRW (jährlich, einwohnerbasiert an Kommunen)",
     "Jährliche Sportpauschale aus dem Gemeindefinanzierungsgesetz (GFG) "
     "an die Kommunen NRWs für Sportstätteninfrastruktur",
     "Breitensport", 69.00, 2024,
     "sportland.nrw – Die Sportpauschale; lsb.nrw"),

    ("Land", "Nordrhein-Westfalen",
     "Staatskanzlei NRW / Landessportbund NRW (LSB NRW)",
     "NRW-Haushalt – LSB-Förderung",
     "Institutionelle Förderung über LSB NRW",
     "Jährliche Förderung des organisierten Sports (Verbände, Vereine, Jugend, "
     "Übungsleiter) über den Landessportbund NRW",
     "Breitensport", 7.56, 2024,
     "land.nrw – mehr als 88 Mio. für Sportvereine NRW"),

    ("Land", "Nordrhein-Westfalen",
     "Staatskanzlei NRW",
     "NRW-Sportmilliarde – Kommunale Sportstätten + Schwimmbäder",
     "NRW-Sportmilliarde: 600 Mio. € kommunale Sportstätten (2025–2030)",
     "Investitionsprogramm für Neubau und Sanierung kommunaler Sportstätten "
     "und Schwimmbäder (600 Mio. €) + Sportpauschale 5 Jahre (375 Mio. €) = "
     "~975 Mio. € Gesamtpaket",
     "Breitensport", 975.00, 2025,
     "sportland.nrw – NRW-Sportmilliarde; lsb.nrw PM (Laufzeit 2025–2030)"),

    ("Land", "Nordrhein-Westfalen",
     "Ministerium für Heimat, Kommunales, Bau und Digitalisierung NRW (MHKBD)",
     "NRW-Haushalt – Investitionspakt Sportstätten",
     "Investitionspakt Sportstätten NRW – 66 Projekte",
     "Gemeinsame Förderung von Bund und Land NRW für 66 Sportstättenprojekte "
     "in Städten und Gemeinden (Bundesanteil + Landesanteil)",
     "Breitensport", 50.00, 2024,
     "mhkbd.nrw Pressemitteilung 2024 – Rund 50 Mio. € für 66 Projekte"),

    ("Land", "Nordrhein-Westfalen",
     "Staatskanzlei NRW",
     "Programm '1.000 x 1.000 – Anerkennung für Sportvereine'",
     "Vereinsanerkennung NRW – '1.000 x 1.000'",
     "Direktzuschüsse an 1.000 ausgewählte Sportvereine (je 1.000 €) als "
     "Anerkennung für ihr ehrenamtliches Engagement",
     "Breitensport", 1.25, 2024,
     "land.nrw Pressemitteilung – Digitalisierungsoffensive / Vereinsförderung"),

    # ── Baden-Württemberg ─────────────────────────────────────
    ("Land", "Baden-Württemberg",
     "Ministerium für Kultus, Jugend und Sport Baden-Württemberg (KM BW)",
     "BW-Haushalt – Sportstättenbauförderung",
     "Sportstättenbau 2024 – 117 kommunale Projekte",
     "Zuschüsse für Neubau und Sanierung von Sporthallen und Freisportanlagen "
     "in Kommunen und Vereinen",
     "Breitensport", 17.30, 2024,
     "km.baden-wuerttemberg.de; Regierungspräsidium Stuttgart PM 2024"),

    ("Land", "Baden-Württemberg",
     "Ministerium für Kultus, Jugend und Sport Baden-Württemberg (KM BW)",
     "BW-Haushalt DH 2025/2026 – Sport",
     "Sportstättenbauförderung BW 2025 (LuKIFG-Nachtrag)",
     "Zusätzliche Sportstättenförderung aus LuKIFG (Landesanteil Sondervermögen): "
     "80 Mio. € als Einmalmaßnahme 2025",
     "Breitensport", 80.00, 2025,
     "fm.baden-wuerttemberg.de Landeshaushalt 2025/2026"),

    ("Land", "Baden-Württemberg",
     "Ministerium für Kultus, Jugend und Sport Baden-Württemberg (KM BW)",
     "BW-Haushalt 2027–2031 – Sportvereine",
     "Langfristpaket Sportvereine BW 2027–2031 (605 Mio. €)",
     "Gesamtpaket für Sportvereine und Verbände in Baden-Württemberg "
     "über fünf Jahre (ab 2027)",
     "Breitensport", 605.00, 2027,
     "stm.bw.de – Mehr als 600 Mio. € für Sportvereine und Verbände"),

    # ── Berlin ────────────────────────────────────────────────
    ("Land", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin",
     "EP 05, Kap. 0510, Tit. 68419",
     "Förderung des Sports (§ 15 SportFG Berlin)",
     "Institutionelle Förderung des organisierten Sports in Berlin nach "
     "§ 15 Sportförderungsgesetz (LSB Berlin, Fachverbände, Vereine)",
     "Spitzensport + Breitensport", 25.00, 2024,
     "berlin.de/sen/inneres Sportförderung; Parlamentsdokument sp19-0093"),

    ("Land", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin",
     "EP 05, Kap. 0510, Tit. 68419 – TA 7",
     "Landestrainerinnen/-trainer Berlin",
     "Vergütung und Betriebskosten der Landestrainer (Nachwuchsleistungssport Berlin)",
     "Spitzensport", 3.65, 2024,
     "Parlamentsdokument sp19-0093; Haushalt Berlin EP 05"),

    ("Land", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin",
     "EP 05, Kap. 0510 – Sportstättensanierung",
     "Sportstättensanierungsprogramm Berlin",
     "Zweckgebundene Mittel für die Sanierung bezirklicher Sportstätten "
     "(2024 abgerufen: 23,12 Mio. €)",
     "Breitensport", 24.15, 2024,
     "berlin.de/rbmskzl Pressemitteilung 2025 – Sportstättensanierung 2024"),

    ("Land", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin",
     "EP 05 – Friedrich-Ludwig-Jahn-Sportpark",
     "Neubau Friedrich-Ludwig-Jahn-Sportpark",
     "Investition in den Neubau des Jahn-Sportparks als modernes Sportzentrum "
     "(Gesamtvolumen Hauptbauabschnitt bis 2025)",
     "Spitzensport + Breitensport", 30.00, 2025,
     "Senatsverwaltung Berlin – Jahn-Sportpark Projekt"),

    ("Land", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin",
     "EP 05, Kap. 0510 – TA 18",
     "'Berlin bewegt sich' – Breitensportbewegung",
     "Kampagnenförderung für Breitensportbewegung und Gesundheitssport",
     "Breitensport", 0.25, 2024,
     "Parlamentsdokument sp19-0093"),

    # ── Hamburg ───────────────────────────────────────────────
    ("Land", "Hamburg",
     "Behörde für Inneres und Sport (BIS) Hamburg",
     "HH-Haushalt – Sportfördervertrag",
     "Sportfördervertrag Hamburg 2025–2026 (mit Hamburger Sportbund HSB)",
     "Neue Sportfördervereinbarung 2025/2026: +750.000 €/Jahr gegenüber Vorperiode; "
     "Steigerung seit 2017/2018 um 46% (9,2 → 13,4 Mio. €/Jahr)",
     "Breitensport", 13.40, 2025,
     "hamburg.de PM – Neuer Sportfördervertrag 2025; SPD Fraktion Hamburg"),

    # ── Sachsen ───────────────────────────────────────────────
    ("Land", "Sachsen",
     "Sächsisches Staatsministerium des Innern (SMI)",
     "Sächsischer Haushalt – Zuwendungsvertrag Sport (LSB Sachsen)",
     "Sportförderung Sachsen 2025 (Zuwendungsvertrag LSB)",
     "Breiten- und Leistungssportförderung inkl. Vereinsförderung (10,3 Mio. €), "
     "Inklusion (136.000 €); Gesamtpaket 2025+2026: 59,7 Mio. €",
     "Spitzensport + Breitensport", 28.80, 2025,
     "sport-fuer-sachsen.de – Neuer Zuwendungsvertrag LSB Sachsen"),

    ("Land", "Sachsen",
     "Sächsisches Staatsministerium des Innern (SMI)",
     "Sächsischer Haushalt – Investive Sportförderung",
     "Investive Sportförderung Sachsen 2026",
     "Investitionen in Sportstätten und Sportinfrastruktur",
     "Breitensport", 0.85, 2026,
     "medienservice.sachsen.de – Investive Sportförderung 2025/2026"),

    # ── Brandenburg ───────────────────────────────────────────
    ("Land", "Brandenburg",
     "Ministerium für Bildung, Jugend und Sport Brandenburg (MBJS)",
     "Brandenburger Haushalt – Sportfördergesetz",
     "Sportförderung Brandenburg 2025 (neues Sportfördergesetz)",
     "Breiten- und Leistungssportförderung nach neuem Sportfördergesetz; "
     "Doppelhaushalt 2025+2026: 55 Mio. €",
     "Spitzensport + Breitensport", 27.00, 2025,
     "LSB Brandenburg – neues Sportfördergesetz; lsb-brandenburg.de Haushaltsplan 2025"),

    ("Land", "Brandenburg",
     "Ministerium für Bildung, Jugend und Sport Brandenburg (MBJS)",
     "Brandenburger Haushalt – Goldener Plan",
     "Goldener Plan Brandenburg (2021–2024, jährlich)",
     "Investitionsprogramm für Sportstätten in Kommunen und Vereinen; "
     "jährlich 6,25 Mio. €",
     "Breitensport", 6.25, 2024,
     "MBJS Brandenburg – Goldener Plan Sportstätten"),

    # ── Rheinland-Pfalz ───────────────────────────────────────
    ("Land", "Rheinland-Pfalz",
     "Ministerium des Innern und für Sport Rheinland-Pfalz (MdI RLP)",
     "RLP-Haushalt – Sportförderung gesamt",
     "Gesamte Sportförderung RLP 2024",
     "Sportstättenförderung (16,7 Mio. €), Projektförderung LSB RLP (13,62 Mio. €), "
     "Leistungssportförderung (3,55 Mio. €), Aus-/Fortbildung (3,6 Mio. €)",
     "Spitzensport + Breitensport", 38.00, 2024,
     "mdi.rlp.de – Rekordförderung Sport RLP"),

    ("Land", "Rheinland-Pfalz",
     "Ministerium des Innern und für Sport Rheinland-Pfalz (MdI RLP)",
     "RLP-Haushalt – Leistungssportförderung",
     "Leistungssportförderung RLP 2025",
     "Steigerung gegenüber 2020 (1,8 Mio. €) um 97% auf 3,55 Mio. € im Jahr 2025",
     "Spitzensport", 3.55, 2025,
     "mdi.rlp.de – Rekordförderung Rheinland-Pfalz"),

    ("Land", "Rheinland-Pfalz",
     "Ministerium des Innern und für Sport Rheinland-Pfalz (MdI RLP)",
     "RLP-Haushalt – LSB RLP Projektförderung",
     "Projektförderung LSB Rheinland-Pfalz 2025",
     "Institutionelle Förderung des Landessportbundes RLP für Breitensport, "
     "Vereine und Verbände",
     "Breitensport", 13.62, 2025,
     "mdi.rlp.de – Rekordförderung Rheinland-Pfalz"),

    # ── Mecklenburg-Vorpommern ────────────────────────────────
    ("Land", "Mecklenburg-Vorpommern",
     "Ministerium für Inneres, Bau und Digitalisierung MV",
     "MV-Haushalt DH 2024/2025 – Sportförderung",
     "Allgemeine Sportförderung MV (gesetzlich festgeschrieben)",
     "Allgemeine Sportförderung gemäß Sportfördergesetz MV; "
     "DH 2024/2025: 12,84 Mio. €; gesetzlich festgeschrieben: 11,92 Mio. €; "
     "Steigerung gegenüber Vorperiode (+3 Mio. €)",
     "Spitzensport + Breitensport", 12.84, 2024,
     "Regierung MV Pressemitteilung; Landtag MV Sozialausschuss 8-919"),

    ("Land", "Mecklenburg-Vorpommern",
     "Ministerium für Inneres, Bau und Digitalisierung MV",
     "EU-ELER-Programm – Sportinfrastruktur MV",
     "EU-ELER-Mittel für Sportinfrastruktur MV",
     "Europäischer Landwirtschaftsfonds für ländliche Entwicklung (ELER): "
     "Förderung von Sportinfrastruktur in ländlichen Gebieten MV bis 2027",
     "Breitensport", 12.90, 2027,
     "Regierung MV – EU-ELER Sportinfrastruktur"),

    # ── Sachsen-Anhalt ────────────────────────────────────────
    ("Land", "Sachsen-Anhalt",
     "Ministerium für Inneres und Sport Sachsen-Anhalt (MI ST)",
     "ST-Haushalt – Vereinssportstättenbau",
     "Vereinssportstättenbau Sachsen-Anhalt 2024",
     "Förderung von Neubau und Sanierung vereinseigener Sportanlagen "
     "(98 geprüfte Bauvorhaben bewilligt)",
     "Breitensport", 5.60, 2024,
     "lsb-sachsen-anhalt.de – Vereinssportstätten 2024 mit 5,6 Mio. € gefördert"),

    # ── Schleswig-Holstein ────────────────────────────────────
    ("Land", "Schleswig-Holstein",
     "Ministerium für Inneres, Kommunales, Wohnen und Sport SH",
     "SH-Haushalt – Sportstättenförderung",
     "Sportstättensanierung SH (jährlich, Schwerpunkt Schwimmsport)",
     "Jährliche Landesförderung der Sportstättensanierung mit Schwerpunkt "
     "auf Schwimmsportanlagen",
     "Breitensport", 2.00, 2024,
     "SPD SH Sportpolitik; SH Haushaltsplan"),

    ("Land", "Schleswig-Holstein",
     "Ministerium für Inneres, Kommunales, Wohnen und Sport SH",
     "SH-Haushalt – Sportförderung Glücksspielmittel",
     "Sportförderung SH aus Lotterieabgabe",
     "Sportförderung über Lottomittel an den Landessportverband SH "
     "(Schätzwert 2024 auf Basis 2015-Wert: 8 Mio. €)",
     "Breitensport", 8.00, 2024,
     "SPD SH – Sportpolitik; Wert Stand 2015"),

    # ── Niedersachsen ─────────────────────────────────────────
    ("Land", "Niedersachsen",
     "Ministerium für Inneres, Sport und Digitalisierung Niedersachsen (MI Nds.)",
     "Nds.-Haushalt – Landeshilfe an LSB (NSportFG)",
     "Landeshilfe Niedersachsen an LSB (NSportFG)",
     "Jährliche Förderung nach Niedersächsischem Sportfördergesetz an den "
     "Landessportbund Niedersachsen für Vereine, Verbände, Übungsleiter",
     "Breitensport", 35.20, 2024,
     "mi.niedersachsen.de Sportbericht 2024"),

    ("Land", "Niedersachsen",
     "Ministerium für Inneres, Sport und Digitalisierung Niedersachsen (MI Nds.)",
     "Nds.-Haushalt – Glücksspielabgabe Sport",
     "Sportförderung Nds. aus Lotterieabgabe",
     "Sportförderung über Glücksspielabgabe (Lottomittel) in Niedersachsen",
     "Breitensport", 14.80, 2024,
     "mi.niedersachsen.de Sportbericht 2024"),

    ("Land", "Niedersachsen",
     "Ministerium für Inneres, Sport und Digitalisierung Niedersachsen (MI Nds.)",
     "Nds.-Haushalt 2025 – Sportstätteninvestitionsprogramm",
     "Sportstätteninvestitionsprogramm Nds. 2025",
     "20 Mio. € kommunaler Sportstättenbau + 5 Mio. € Vereinssportstättenbau",
     "Breitensport", 25.00, 2025,
     "mi.niedersachsen.de – 25 Mio. Euro für Sportstätten in Niedersachsen"),

    # ── Hessen ────────────────────────────────────────────────
    ("Land", "Hessen",
     "Hessisches Ministerium für Familie, Senioren, Sport, Gesundheit und Pflege (HMFG)",
     "HE-Haushalt – Sportprojektförderung 2024",
     "Sportprojektförderung Hessen 2024 (1.001 Projekte)",
     "Förderung von 1.001 Sportprojekten in Hessen im Jahr 2024 "
     "(Q3-Stand: 233 Projekte mit 10,6 Mio. €; Q1: 3,6 Mio. €)",
     "Breitensport", 40.00, 2024,
     "familie.hessen.de – 233 Sportprojekte 10,6 Mio. €; innen.hessen.de"),

    ("Land", "Hessen",
     "Hessisches Ministerium des Innern und für Sport (HMdIS)",
     "HE-Haushalt – LSB Hessen institutionelle Förderung",
     "Institutionelle Förderung LSB Hessen",
     "Regelzuwendung an den Landessportbund Hessen für institutionelle Aufgaben",
     "Breitensport", 1.30, 2024,
     "innen.hessen.de – Landessportbund erhält Förderung von rund 1,3 Mio. €"),

    # ── Bremen ────────────────────────────────────────────────
    ("Land", "Bremen",
     "Senator für Inneres / Sportdeputation Bremen",
     "Bremer Haushalt DH 2024/2025 – Sport",
     "Gesamtsportförderung Bremen DH 2024/2025",
     "Gesamte Sportförderung inkl. Großprojekt Westbad (11,03 Mio. € 2024, "
     "4,51 Mio. € 2025) und Vereinssportstättensanierung (>800.000 €)",
     "Breitensport", 32.57, 2024,
     "senatspressestelle.bremen.de – Sportdeputation Haushalt 2024/2025"),

    # ── Saarland ──────────────────────────────────────────────
    ("Land", "Saarland",
     "Ministerium für Inneres, Bauen und Sport Saarland (MIBS)",
     "SL-Haushalt – Spitzensportförderung",
     "Zuwendungsbudget Spitzensport Saarland 2024 (erhöht)",
     "Erhöhung um 250.000 € auf 332.800 € Gesamtvolumen für den Spitzensport; "
     "zusätzlich 600.000 € Sportoto-Mittel (Olympia Paris 2024)",
     "Spitzensport", 0.33, 2024,
     "saarland.de – Olympia 2024 Zuwendungen; Homburg1.de PM"),

    ("Land", "Saarland",
     "Ministerium für Inneres, Bauen und Sport Saarland (MIBS)",
     "SL-Haushalt – Sportveranstaltungen",
     "Budget Sportveranstaltungen Saarland 2024 (erhöht)",
     "Förderbudget für besondere sportliche Veranstaltungen mit überregionalem "
     "Stellenwert (erhöht um 190.000 € auf 710.000 €)",
     "Spitzensport + Breitensport", 0.71, 2024,
     "Homburg1.de PM 2024 – Olympia-Zuwendungen Saarland"),

    # ── Thüringen ─────────────────────────────────────────────
    ("Land", "Thüringen",
     "Thüringer Ministerium für Bildung, Jugend und Sport (TMBJS)",
     "TH-Haushalt – Vereinseigener Sportstättenbau",
     "Vereinseigener Sportstättenbau Thüringen",
     "Investitionszuschüsse für Bau und Sanierung vereinseigener Sportanlagen "
     "(LSB Thüringen verwaltet)",
     "Breitensport", 2.00, 2024,
     "thueringen-sport.de – 2 Mio. für vereinseigene Sportstätten 2024"),

    ("Land", "Thüringen",
     "Thüringer Ministerium für Bildung, Jugend und Sport (TMBJS)",
     "TH-Haushalt DH 2026/2027 – Gesamtsportförderung",
     "Gesamtförderung Sport Thüringen 2026/2027",
     "Gemeindesportstätten: 12,75 Mio. € (2026) / 10,55 Mio. €; "
     "Nachwuchstrainer: 3,25 Mio. € (2026); Spitzensport: ~1,1 Mio. €; "
     "Vereine/Verbände: min. 2,7 Mio. €; Gesamtschätzung ~35 Mio. €/Jahr",
     "Spitzensport + Breitensport", 35.00, 2026,
     "TMBJS Thüringen – Haushaltsplanung 2026/2027; LSB Thüringen"),

    # ══════════════════════════════════════════════════════════
    # KOMMUNALE EBENE
    # ══════════════════════════════════════════════════════════

    ("Kommune", "Bayern",
     "Landeshauptstadt München – Referat für Bildung und Sport",
     "Stadthaushalt München – Sportbetriebspauschale + Unterhalt",
     "Direkte Sportförderung München 2025 (Vereine + Unterhalt)",
     "Sportbetriebspauschale und Unterhaltsförderung für Vereinssportanlagen; "
     "+500.000 € Erhöhung beider Positionen 2025 gegenüber Vorjahr",
     "Breitensport", 7.00, 2025,
     "spd-muenchen.de – Stadt unterstützt Sportvereine 2025 mit rund 7 Mio. €"),

    ("Kommune", "Berlin",
     "Senatsverwaltung für Inneres und Sport Berlin – Bezirke",
     "Berliner Bezirkshaushalte – Sportstättensanierung",
     "Sportstättensanierungsprogramm Berlin an Bezirke 2024",
     "Zweckgebundene Landesmittel für Sanierung bezirklicher Sportstätten "
     "(abgerufen 2024: 23,12 Mio. €; Programmmittel: 24,15 Mio. €)",
     "Breitensport", 23.12, 2024,
     "berlin.de/rbmskzl – Sportstättensanierung Bericht 2024"),

    ("Kommune", "Nordrhein-Westfalen",
     "Stadt Köln – Dezernat für Sport und Infrastruktur",
     "Kölner Stadthaushalt DH 2023/2024 – Sport",
     "Sportförderung und Unterhaltung Sportstätten Köln",
     "Maßnahmenkatalog Sportförderung und Unterhaltung kommunaler Sportstätten "
     "im Doppelhaushalt 2023/2024 (ca. 22,9 Mio. €/Jahr)",
     "Breitensport", 22.90, 2024,
     "gruenekoeln.de – Analyse Kölner Haushaltssatzung 2023/2024"),

    ("Kommune", "Hamburg",
     "Bezirksämter Hamburg / Behörde für Inneres und Sport",
     "Hamburger Haushalt – Kommunale Sportförderung",
     "Kommunale Sportförderung Hamburg (Bezirke + BIS)",
     "Sportförderung durch die Hamburgischen Bezirksämter und die Behörde "
     "für Inneres und Sport (Sportstättenbetrieb, Vereine, Veranstaltungen)",
     "Breitensport", 20.00, 2024,
     "BIS Hamburg; Hamburger Haushalt 2024 (Schätzwert)"),

    ("Kommune", "Nordrhein-Westfalen",
     "Stadt Düsseldorf – Sportamt",
     "Düsseldorfer Stadthaushalt – Sport",
     "Kommunale Sportförderung Düsseldorf",
     "Betrieb kommunaler Sportanlagen und Vereinsförderung "
     "(inkl. Hallennutzung, Zuschüsse)",
     "Breitensport", 18.00, 2024,
     "Stadthaushalt Düsseldorf 2024 (Schätzwert; Sportamt Düsseldorf)"),

    ("Kommune", "Bayern",
     "Stadt Nürnberg – Stadtrat Sport",
     "Nürnberger Stadthaushalt – Sport",
     "Kommunale Sportförderung Nürnberg",
     "Betrieb kommunaler Sportstätten und Vereinsförderung in Nürnberg",
     "Breitensport", 12.00, 2024,
     "Stadthaushalt Nürnberg 2024 (Schätzwert)"),

    ("Kommune", "Sachsen",
     "Landeshauptstadt Dresden – Sportamt",
     "Dresdner Stadthaushalt – Sport",
     "Kommunale Sportförderung Dresden",
     "Betrieb kommunaler Sportstätten, Vereinsförderung und Veranstaltungen",
     "Breitensport", 15.00, 2024,
     "Stadthaushalt Dresden 2024 (Schätzwert; Sportamt Dresden)"),
]

# ─── Spaltentitel ──────────────────────────────────────────────────────────────
COLS = [
    "Ebene",
    "Bundesland",
    "Träger / Ministerium",
    "Einzelplan / Kapitel / Titel",
    "Maßnahme / Titel",
    "Beschreibung",
    "Förderbereich",
    "Betrag (Mio. €)",
    "Haushaltsjahr",
    "Quelle / Anmerkung",
]

# ─── Hilfsfunktionen ──────────────────────────────────────────────────────────

def hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row: int, values: list, fill: str,
               font_color: str = "FFFFFF", size: int = 10):
    f = hfill(fill)
    ft = Font(bold=True, color=font_color, size=size)
    brd = thin_border()
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = f
        cell.font = ft
        cell.border = brd
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")

def data_row(ws, row: int, values: list, bg: str | None = None):
    brd = thin_border()
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        if bg:
            cell.fill = hfill(bg)
        cell.border = brd
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 8:  # Betrag
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="top")

def set_col_widths(ws, widths: list[tuple[str, float]]):
    for col, w in widths:
        ws.column_dimensions[col].width = w

# ─── DATEI 1: Nach Haushaltspositionen ────────────────────────────────────────

def file_positionen():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alle Positionen"

    # Titelzeile
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = ("Sportförderung in Deutschland – Bundesministerien, Landesministerien "
               "und Kommunen │ Haushaltspositionen 2024/2025/2026")
    c.font = Font(bold=True, size=14, color="1A237E")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 44

    # Hinweiszeile
    ws.merge_cells("A2:J2")
    h = ws["A2"]
    h.value = ("Beträge in Mio. €  │  Quellen: BMI EP 06, Bundestag hib, DOSB, Landeshaushalte, "
               "Landessportbünde, kommunale Haushalte  │  Kommunale Werte teilweise Schätzungen "
               "│  Stand: März 2026")
    h.font = Font(italic=True, size=8, color="555555")
    h.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28

    header_row(ws, 3, COLS, "1A237E")
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"

    EBENE_BG = {"Bund": "E3F2FD", "Land": "E8F5E9", "Kommune": "FFF8E1"}

    sorted_data = sorted(DATA, key=lambda r: (
        {"Bund": 0, "Land": 1, "Kommune": 2}.get(r[0], 9), r[1], r[4]
    ))

    for i, row in enumerate(sorted_data, start=4):
        data_row(ws, i, list(row), EBENE_BG.get(row[0]))

    # Summe gesamt
    last = 3 + len(sorted_data)
    ws.cell(row=last + 2, column=7, value="GESAMTSUMME (Mio. €)").font = Font(bold=True, size=11)
    sc = ws.cell(row=last + 2, column=8, value=round(sum(r[7] for r in DATA), 2))
    sc.number_format = '#,##0.00'
    sc.font = Font(bold=True, size=11, color="B71C1C")
    sc.alignment = Alignment(horizontal="right")
    ws.cell(row=last + 3, column=7,
            value="Hinweis: Mehrfachzählungen möglich (Bund → Land → Kommune-Flüsse)").font = Font(
        italic=True, size=8, color="777777")

    col_widths = [("A",13),("B",24),("C",48),("D",34),("E",44),
                  ("F",62),("G",24),("H",14),("I",13),("J",56)]
    set_col_widths(ws, col_widths)

    # Zusatzblätter per Ebene
    for ebene, fill_h, fill_r, title_suffix in [
        ("Bund",    "0D47A1", "DCEEFB", "Bundesebene"),
        ("Land",    "1B5E20", "DCF5DC", "Länderebene"),
        ("Kommune", "E65100", "FFF3E0", "Kommunale Ebene"),
    ]:
        ws2 = wb.create_sheet(title_suffix)
        header_row(ws2, 1, COLS, fill_h)
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"
        rows = [r for r in sorted_data if r[0] == ebene]
        for i, row in enumerate(rows, start=2):
            data_row(ws2, i, list(row), fill_r)
        # Summe
        last2 = 1 + len(rows)
        ws2.cell(row=last2 + 1, column=7, value=f"Summe {title_suffix}").font = Font(bold=True)
        sc2 = ws2.cell(row=last2 + 1, column=8,
                        value=round(sum(r[7] for r in rows), 2))
        sc2.number_format = '#,##0.00'
        sc2.font = Font(bold=True, color="B71C1C")
        sc2.alignment = Alignment(horizontal="right")
        set_col_widths(ws2, col_widths)

    path = "/home/user/claude-plugins-journalism/sportfoerderung_nach_haushaltspositionen.xlsx"
    wb.save(path)
    print(f"Datei 1 gespeichert: {path}")
    return path


# ─── DATEI 2: Nach Bundesland ─────────────────────────────────────────────────

BL_ORDER = [
    "Gesamt (Bund)",
    "Baden-Württemberg", "Bayern", "Berlin", "Brandenburg", "Bremen",
    "Hamburg", "Hessen", "Mecklenburg-Vorpommern", "Niedersachsen",
    "Nordrhein-Westfalen", "Rheinland-Pfalz", "Saarland", "Sachsen",
    "Sachsen-Anhalt", "Schleswig-Holstein", "Thüringen",
]

BL_FILL = {
    "Gesamt (Bund)":           ("1A237E", "FFFFFF"),
    "Baden-Württemberg":       ("006400", "FFFFFF"),
    "Bayern":                  ("0055A4", "FFFFFF"),
    "Berlin":                  ("CC0000", "FFFFFF"),
    "Brandenburg":             ("C8102E", "FFFFFF"),
    "Bremen":                  ("002868", "FFFFFF"),
    "Hamburg":                 ("E2001A", "FFFFFF"),
    "Hessen":                  ("CC0000", "FFFFFF"),
    "Mecklenburg-Vorpommern":  ("006AB3", "FFFFFF"),
    "Niedersachsen":           ("FFCC00", "000000"),
    "Nordrhein-Westfalen":     ("009A44", "FFFFFF"),
    "Rheinland-Pfalz":         ("C8102E", "FFFFFF"),
    "Saarland":                ("003087", "FFFFFF"),
    "Sachsen":                 ("006600", "FFFFFF"),
    "Sachsen-Anhalt":          ("FFFF00", "000000"),
    "Schleswig-Holstein":      ("003087", "FFFFFF"),
    "Thüringen":               ("CC0000", "FFFFFF"),
}

BL_ROW_BG = {
    "Gesamt (Bund)":           "D9E8FB",
    "Baden-Württemberg":       "E8F5E8",
    "Bayern":                  "E8F0FB",
    "Berlin":                  "FFECEC",
    "Brandenburg":             "FFECEC",
    "Bremen":                  "ECF0FF",
    "Hamburg":                 "FFECEC",
    "Hessen":                  "FFECEC",
    "Mecklenburg-Vorpommern":  "E8F0FF",
    "Niedersachsen":           "FFFDE8",
    "Nordrhein-Westfalen":     "ECFBEC",
    "Rheinland-Pfalz":         "FFECEC",
    "Saarland":                "ECF0FF",
    "Sachsen":                 "ECFBEC",
    "Sachsen-Anhalt":          "FFFFCC",
    "Schleswig-Holstein":      "ECF0FF",
    "Thüringen":               "FFECEC",
}


def file_bundesland():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nach Bundesland"

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = ("Sportförderung in Deutschland – Übersicht nach Bundesland │ "
               "Bundesministerien, Landesministerien und Kommunen │ 2024/2025/2026")
    c.font = Font(bold=True, size=14, color="1A237E")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:J2")
    h = ws["A2"]
    h.value = ("Beträge in Mio. €  │  Mehrfachzählungen möglich  │  "
               "Kommunale Werte = Schätzungen  │  Stand: März 2026")
    h.font = Font(italic=True, size=8, color="555555")
    h.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 24

    header_row(ws, 3, COLS, "1A237E")
    ws.row_dimensions[3].height = 32
    ws.freeze_panes = "A4"

    current = 4
    sorted_data = sorted(DATA, key=lambda r: (
        BL_ORDER.index(r[1]) if r[1] in BL_ORDER else 99, r[0], r[4]
    ))

    bl_totals = {}
    for bl in BL_ORDER:
        rows = [r for r in sorted_data if r[1] == bl]
        if not rows:
            continue

        # Bundesland-Header
        ws.merge_cells(start_row=current, start_column=1,
                       end_row=current, end_column=10)
        hc = ws.cell(row=current, column=1,
                     value=f"  ▶  {bl.upper()}")
        fill_h, font_c = BL_FILL.get(bl, ("444444", "FFFFFF"))
        hc.fill = hfill(fill_h)
        hc.font = Font(bold=True, size=11, color=font_c)
        hc.alignment = Alignment(vertical="center")
        ws.row_dimensions[current].height = 22
        current += 1

        row_bg = BL_ROW_BG.get(bl, "F9F9F9")
        bl_sum = 0.0
        for row in rows:
            data_row(ws, current, list(row), row_bg)
            bl_sum += row[7]
            current += 1

        # Subtotal
        ws.cell(row=current, column=7,
                value=f"Summe {bl}").font = Font(bold=True)
        ws.cell(row=current, column=7).alignment = Alignment(horizontal="right")
        sc = ws.cell(row=current, column=8, value=round(bl_sum, 2))
        sc.number_format = '#,##0.00'
        sc.font = Font(bold=True)
        sc.fill = hfill("F0F0F0")
        sc.alignment = Alignment(horizontal="right")
        ws.cell(row=current, column=7).fill = hfill("F0F0F0")
        bl_totals[bl] = bl_sum
        current += 2  # Leerzeile

    # Gesamtsumme
    ws.cell(row=current, column=7,
            value="GESAMTSUMME (alle Bundesländer + Bund)").font = Font(bold=True, size=11)
    gtc = ws.cell(row=current, column=8, value=round(sum(bl_totals.values()), 2))
    gtc.number_format = '#,##0.00'
    gtc.font = Font(bold=True, size=11, color="B71C1C")
    gtc.alignment = Alignment(horizontal="right")

    col_widths = [("A",13),("B",24),("C",48),("D",34),("E",44),
                  ("F",62),("G",24),("H",14),("I",13),("J",56)]
    set_col_widths(ws, col_widths)

    # Zusammenfassungsblatt
    ws2 = wb.create_sheet("Zusammenfassung je Bundesland")
    sum_cols = ["Bundesland", "Ebene(n)", "Anzahl Positionen",
                "Summe Förderung (Mio. €)", "Haushaltsjahre", "Anmerkung"]
    header_row(ws2, 1, sum_cols, "1A237E")
    ws2.freeze_panes = "A2"

    for r_idx, bl in enumerate(BL_ORDER, start=2):
        rows = [r for r in DATA if r[1] == bl]
        if not rows:
            continue
        fill_h, font_c = BL_FILL.get(bl, ("EEEEEE", "000000"))
        ebenen = ", ".join(sorted(set(r[0] for r in rows)))
        years = ", ".join(sorted(set(str(r[8]) for r in rows)))
        total = round(sum(r[7] for r in rows), 2)
        note = (f"{len(rows)} Haushaltspositionen; davon "
                f"Spitzensport: {sum(1 for r in rows if 'Spitzensport' in r[6])}; "
                f"Breitensport: {sum(1 for r in rows if 'Breitensport' in r[6])}")
        vals = [bl, ebenen, len(rows), total, years, note]
        for c_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = hfill(fill_h)
            cell.font = Font(color=font_c)
            cell.border = thin_border()
            if c_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    set_col_widths(ws2, [("A",28),("B",22),("C",22),("D",24),("E",22),("F",58)])

    # Balkendiagramm-Datenblatt (Rohdaten für Diagramme)
    ws3 = wb.create_sheet("Diagrammdaten")
    header_row(ws3, 1,
               ["Bundesland", "Bund-Anteil (Mio. €)", "Land-Anteil (Mio. €)",
                "Kommunal-Anteil (Mio. €)", "Gesamt (Mio. €)",
                "Spitzensport-Anteil (Mio. €)", "Breitensport-Anteil (Mio. €)"],
               "37474F")
    ws3.freeze_panes = "A2"
    for r_idx, bl in enumerate(BL_ORDER, start=2):
        rows = [r for r in DATA if r[1] == bl]
        if not rows:
            continue
        bund = round(sum(r[7] for r in rows if r[0] == "Bund"), 2)
        land = round(sum(r[7] for r in rows if r[0] == "Land"), 2)
        komm = round(sum(r[7] for r in rows if r[0] == "Kommune"), 2)
        spitz = round(sum(r[7] for r in rows if "Spitzensport" in r[6] and "Breiten" not in r[6]), 2)
        breit = round(sum(r[7] for r in rows if "Breitensport" in r[6] and "Spitz" not in r[6]), 2)
        ws3.cell(row=r_idx, column=1, value=bl)
        for c_idx, val in enumerate([bund, land, komm, bund+land+komm, spitz, breit], 2):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
    set_col_widths(ws3, [("A",28),("B",22),("C",22),("D",22),("E",18),("F",24),("G",24)])

    path = "/home/user/claude-plugins-journalism/sportfoerderung_nach_bundeslaendern.xlsx"
    wb.save(path)
    print(f"Datei 2 gespeichert: {path}")
    return path


if __name__ == "__main__":
    p1 = file_positionen()
    p2 = file_bundesland()
    total = sum(r[7] for r in DATA)
    bund_total = sum(r[7] for r in DATA if r[0] == "Bund")
    land_total = sum(r[7] for r in DATA if r[0] == "Land")
    komm_total = sum(r[7] for r in DATA if r[0] == "Kommune")
    print(f"\nStatistik:")
    print(f"  Gesamtpositionen: {len(DATA)}")
    print(f"  Summe gesamt:     {total:,.2f} Mio. €")
    print(f"  davon Bund:       {bund_total:,.2f} Mio. €")
    print(f"  davon Länder:     {land_total:,.2f} Mio. €")
    print(f"  davon Kommunen:   {komm_total:,.2f} Mio. €")
    print(f"\nDateien:")
    print(f"  1. {p1}")
    print(f"  2. {p2}")
